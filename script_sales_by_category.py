import time
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ CONFIG ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
# DRIVER_PATH      = r"C:\Users\PC\Downloads\edgedriver_win64\msedgedriver.exe"
DRIVER_PATH = "msedgedriver.exe"
DEBUGGER_ADDRESS = "127.0.0.1:9222"
MASTER_XLSX      = r"D:\Kaam\Muneef 0\Test Automation\Weekly Sales summary (week 37).xlsx"
CURRENT_WEEK     = "Sep 08 - Sep 14"  # exactly as it appears in column A

# Map the exact dropdown label text ‚Üí sheet name
dropdown_to_sheet = {
    "Karachi Kabab Wala - Queen Street":    "Kababwala - Queen",
    "Pizza Karachi- Eglinton":      "Pizza K Eglinton",
    "Pizza Karachi -Heartland":     "Pizza K Heartland",
    "Karachi Kabab Wala":           "Kababwala",
    "Karachi Food Court":           "Karachi Food Court",
    "Pizza Karachi Downtown TO":    "Queen St.",
    "Pizza Karachi- Highway Karahi":"Highway",
    "Pizza Karachi - Wonderland":   "Jane",
    "Pizza Karachi - Lebovic":      "Lebovic",
    "Pizza Karachi - Ajax":         "Ajax",
    "Pizza Karachi - Markham Rd":   "Markham",
}

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ HELPERS ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
def parse_currency(txt: str) -> float:
    """Convert something like '$1,234.56' into 1234.56"""
    return float(txt.replace('$', '').replace(',', '').strip())

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ SETUP SELENIUM ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ LOAD EXCEL ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
wb = load_workbook(MASTER_XLSX)

try:
    # Open dropdown once
    dd_button = wait.until(EC.element_to_be_clickable((
        By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
    )))
    dd_button.click()
    time.sleep(0.5)

    for label_text, sheet_name in dropdown_to_sheet.items():
        print(f"\n‚ñ∂Ô∏è Processing ‚Äú{label_text}‚Äù ‚Üí sheet ‚Äú{sheet_name}‚Äù")

        # Select restaurant from dropdown
        xpath_label = (
            f"//div[@data-pw='radio-dropdown-selector']"
            f"//label[.//span[text()='{label_text}']]"
        )
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath_label))).click()

        # Wait for table to load
        wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"] tbody'
        )))
        time.sleep(0.5)

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ SCRAPE HEADER TO FIND ‚ÄúNet Sales‚Äù INDEX ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        header_cells = driver.find_elements(
            By.CSS_SELECTOR,
            'table[data-pw="report-table-data"] thead tr th'
        )
        header_texts = [
            th.text.strip().replace('\n', ' ')
            for th in header_cells
        ]
        try:
            net_sales_idx = header_texts.index("Net Sales")
        except ValueError:
            raise RuntimeError("Could not find ‚ÄúNet Sales‚Äù column in the table header")

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ PREPARE EXCEL SHEET & HEADER MAPPING ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        ws = wb[sheet_name]

        # Find target row for CURRENT_WEEK in column A
        target_row = None
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value == CURRENT_WEEK:
                target_row = row[0].row
                break
        if not target_row:
            raise RuntimeError(f"Week '{CURRENT_WEEK}' not found in sheet '{sheet_name}'")

        # Read headers from column M onward until "Net Sales"
        excel_header_to_col = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=13, max_col=ws.max_column):
            for cell in row:
                header = cell.value
                if header is None:
                    continue
                excel_header_to_col[header.strip()] = cell.column
    
                if header.strip() == "Net Sales":
                    # Stop once we find the "Net Sales" column
                    break
            if "Net Sales" in excel_header_to_col:
                break

        if "Net Sales" not in excel_header_to_col:
            raise RuntimeError(f"'Net Sales' header not found in sheet '{sheet_name}'")

        # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ SCRAPE AND WRITE EACH ROW ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
        tbody = driver.find_element(
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"] tbody'
        )
        rows = tbody.find_elements(By.TAG_NAME, "tr")

        for tr in rows:
            # Category name from first cell
            cat_cell = tr.find_element(By.XPATH, "./td[1]")
            category = cat_cell.text.strip()

            # Net Sales from the correct column
            net_txt = tr.find_element(
                By.XPATH,
                f"./td[{net_sales_idx + 1}]"
            ).text.strip()
            try:
                net_val = parse_currency(net_txt)
            except Exception:
                print(f"‚ö†Ô∏è Warning: could not parse Net Sales '{net_txt}' for '{category}'. Skipping.")
                continue

            # If this is the summary row, write into the Net Sales column
            if category.lower().startswith("report summary"):
                summary_col = excel_header_to_col["Net Sales"]
                ws.cell(row=target_row, column=summary_col, value=net_val)
                print(f"   üìä Summary Net Sales = {net_val} ‚Üí column {summary_col}")
                continue

            # Otherwise, match category to one of the Excel headers
            # e.g. "Food" ‚Üí header "Food (Pizza K)"
            match_header = None
            for hdr in excel_header_to_col:
                if hdr.lower().startswith(category.lower()):
                    match_header = hdr
                    break

            if not match_header:
                print(f"‚ö†Ô∏è Category '{category}' not found in sheet '{sheet_name}'.")
                continue

            col_idx = excel_header_to_col[match_header]
            ws.cell(row=target_row, column=col_idx, value=net_val)
            print(f"   üè∑ {category!r} = {net_val} ‚Üí '{match_header}' (col {col_idx})")

        # Re-open dropdown for next iteration
        dd_button = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
        )))
        dd_button.click()
        time.sleep(0.5)

    # ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ SAVE WORKBOOK ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    wb.save(MASTER_XLSX)
    print("\n‚úÖ All restaurants processed and master workbook updated!")

finally:
    driver.quit()
