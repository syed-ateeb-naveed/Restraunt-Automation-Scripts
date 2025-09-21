import time, glob, os, logging
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# ───── CONFIG ─────
DRIVER_PATH     = "msedgedriver.exe"
DEBUGGER_ADDR   = "127.0.0.1:9222"
SUMMARY_FILE    = r"D:\Kaam\Muneef 0\Test Automation\Weekly Sales summary (week 37).xlsx"
CURRENT_WEEK    = "Sep 08 - Sep 14"  # exactly as in column A

dropdown_to_sheet = {
    "Karachi Kabab Wala - Queen Street":    "Kababwala - Queen",
    "Pizza Karachi- Eglinton":       "Pizza K Eglinton",
    "Pizza Karachi -Heartland":      "Pizza K Heartland",
    "Karachi Kabab Wala":            "Kababwala",
    "Karachi Food Court":            "Karachi Food Court",
    "Pizza Karachi Downtown TO":     "Queen St.",
    "Pizza Karachi- Highway Karahi": "Highway",
    "Pizza Karachi - Wonderland":    "Jane",
    "Pizza Karachi - Lebovic":       "Lebovic",
    "Pizza Karachi - Ajax":          "Ajax",
    "Pizza Karachi - Markham Rd":    "Markham",
}

BUCKETS = [
    ("8AM - 11AM",   ["8am - 9am","9am - 10am","10am - 11am"]),
    ("11AM - 3PM",   ["11am - 12pm","12pm - 1pm","1pm - 2pm","2pm - 3pm"]),
    ("3PM - 6PM",    ["3pm - 4pm","4pm - 5pm","5pm - 6pm"]),
    ("6PM - 11PM",   ["6pm - 7pm","7pm - 8pm","8pm - 9pm","9pm - 10pm","10pm - 11pm"]),
    ("11PM - 7AM",   ["11pm - 12am","12am - 1am","1am - 2am","2am - 3am","3am - 4am","4am - 5am","5am - 6am","6am - 7am"]),
]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ───── SELENIUM SETUP ─────
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDR)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

try:
    for label, sheet_name in dropdown_to_sheet.items():
        logging.info(f"Processing {label} → sheet {sheet_name}")

        # 1) Open dropdown & select restaurant
        dd = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,"button[data-testid='radio-dropdown-selector-input']")))
        dd.click(); time.sleep(0.5)
        xpath = f"//div[@data-pw='radio-dropdown-selector']//label[.//span[text()='{label}']]"
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        time.sleep(1)

        # 2) Switch to Hourly Sales
        wait.until(EC.element_to_be_clickable((By.XPATH,"//button[@data-pw='hourly-tab-tab']"))).click()
        time.sleep(0.5)
        # 3) Select Bill Start time
        wait.until(EC.element_to_be_clickable((By.XPATH,"//label[@data-pw='bill-start-time']"))).click()
        time.sleep(0.5)
        # 4) Wait for table
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,'table[data-pw="report-table-data"] tbody')))
        time.sleep(0.5)

        # 5) Scrape hour‐of‐day rows → sum across days
        tbody = driver.find_element(By.CSS_SELECTOR,'table[data-pw="report-table-data"] tbody')
        hour_sums = {}
        for tr in tbody.find_elements(By.TAG_NAME,"tr"):
            lbl = tr.find_element(By.XPATH,"./td[1]").text.strip().lower()
            if lbl.startswith("report summary"): continue
            vals = tr.find_elements(By.XPATH,"./td[position()>1]")
            total = sum(float(td.text.replace("$","").replace(",","") or 0) for td in vals)
            hour_sums[lbl] = total

        # 6) Compute bucket values, missing treated as 0
        buckets = []
        for _, hours in BUCKETS:
            s=0.0
            for h in hours:
                if h not in hour_sums:
                    logging.warning(f" Missing '{h}' for {sheet_name}, using 0")
                s += hour_sums.get(h,0.0)
            buckets.append(s)

        # 7) Open workbook & locate sheet
        wb=load_workbook(SUMMARY_FILE)
        if sheet_name not in wb.sheetnames:
            logging.error(f"Sheet not found: {sheet_name}")
            continue
        ws=wb[sheet_name]

        # 8) Locate merged header “Daily and Hourly Sales (TBD)” in row1
        merged_range=None
        for m in ws.merged_cells.ranges:
            if m.min_row==1 and ws.cell(1,m.min_col).value=="Daily and Hourly Sales (TBD)":
                merged_range=m; break
        if not merged_range:
            logging.error("Header not found in row 1"); continue
        start_col=merged_range.min_col

        # 9) Find row matching CURRENT_WEEK in column A
        target_row=None
        for r in range(2,ws.max_row+1):
            if str(ws.cell(r,1).value).strip()==CURRENT_WEEK:
                target_row=r; break
        if not target_row:
            logging.error(f"Week '{CURRENT_WEEK}' not found in {sheet_name}"); continue

        # 10) Write bucket sums into that row under the bucket columns
        for i,val in enumerate(buckets):
            ws.cell(row=target_row, column=start_col+i, value=val)

        wb.save(SUMMARY_FILE)
        logging.info(f" ✔ Saved {sheet_name}")

finally:
    driver.quit()
