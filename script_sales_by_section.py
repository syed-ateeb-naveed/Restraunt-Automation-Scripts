import time
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# ───── CONFIG ─────
# DRIVER_PATH      = r"C:\Users\PC\Downloads\edgedriver_win64\msedgedriver.exe"
DRIVER_PATH = "msedgedriver.exe"
DEBUGGER_ADDRESS = "127.0.0.1:9222"
MASTER_XLSX      = r"D:\Kaam\Muneef 0\Test Automation\Weekly Sales summary (week 37).xlsx"
CURRENT_WEEK     = "Sep 08 - Sep 14"  # exactly as in column A

# Map the exact dropdown label text → sheet name
dropdown_to_sheet = {
    "Karachi Kabab Wala - Queen Street":    "Kababwala - Queen",
    "Pizza Karachi- Eglinton":      "Pizza K Eglinton",
    "Pizza Karachi -Heartland":     "Pizza K Heartland",
    "Karachi Kabab Wala":           "Kababwala",
    "Karachi Food Court":           "Karachi Food Court",
    "Pizza Karachi Downtown TO":    "Queen St.",
    "Pizza Karachi- Highway Karahi":"Highway",
    "Pizza Karachi - Wonderland":   "Jane",
    "Pizza Karachi - Lebovic":      "Lebovic",
    "Pizza Karachi - Ajax":         "Ajax",
    "Pizza Karachi - Markham Rd":   "Markham",
}

# ───── HELPERS ─────
def parse_currency(txt: str) -> float:
    return float(txt.replace('$','').replace(',','').strip())

# ───── SETUP SELENIUM ─────
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

# ───── LOAD EXCEL ─────
wb = load_workbook(MASTER_XLSX)

try:
    # 1) Open the dropdown once
    dd_button = wait.until(EC.element_to_be_clickable((
        By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
    )))
    dd_button.click()
    time.sleep(0.5)

    for label_text, sheet_name in dropdown_to_sheet.items():
        print(f"\n▶️ Processing: “{label_text}” → sheet “{sheet_name}”")

        # 2) Select its radio button
        xpath_label = (
            f"//div[@data-pw='radio-dropdown-selector']"
            f"//label[.//span[text()='{label_text}']]"
        )
        label = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_label)))
        label.click()

        # 3) Wait for table to reload under new restaurant
        wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"] tbody'
        )))
        time.sleep(0.5)

        # 4) Scrape the Sales-By-Sections rows
        tbody = driver.find_element(
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"] tbody'
        )
        rows  = tbody.find_elements(By.TAG_NAME, "tr")

        remaining = skip_val = dd_val = uber_val = 0.0

        for tr in rows:
            sec = tr.find_element(By.XPATH, "./td[1]").text.strip()
            net_txt = tr.find_element(By.XPATH, "./td[last()]").text.strip()
            try:
                net = parse_currency(net_txt)
            except:
                continue
            key = sec.lower()
            if key.startswith("report summary"):
                continue
            if key == "skip the dishes":
                skip_val = net
            elif key == "doordash":
                dd_val = net
            elif key == "ubereats":
                uber_val = net
            else:
                remaining += net

        total = remaining + skip_val + dd_val + uber_val
        print(f"    Remaining={remaining}, Skip={skip_val}, DoorDash={dd_val}, Uber={uber_val}, Total={total}")

        # 5) Write into master Excel (cols B–F = 2–6)
        ws = wb[sheet_name]
        target_row = None
        for row in ws.iter_rows(min_row=2, max_col=1):
            if row[0].value == CURRENT_WEEK:
                target_row = row[0].row
                break
        if not target_row:
            raise RuntimeError(f"Week '{CURRENT_WEEK}' not found in sheet '{sheet_name}'")

        ws.cell(row=target_row, column=2, value=remaining)
        ws.cell(row=target_row, column=3, value=skip_val)
        ws.cell(row=target_row, column=4, value=dd_val)
        ws.cell(row=target_row, column=5, value=uber_val)
        ws.cell(row=target_row, column=6, value=total)

        # 6) Re-open dropdown for next restaurant
        dd_button = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
        )))
        dd_button.click()
        time.sleep(0.5)

    # 7) Save all updates
    wb.save(MASTER_XLSX)
    print("\n✅ All restaurants processed and master workbook updated!")

finally:
    driver.quit()
