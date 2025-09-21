import time
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# ───── CONFIG ─────
# DRIVER_PATH      = r"C:\Users\PC\Downloads\edgedriver_win64\msedgedriver.exe"
DRIVER_PATH = "msedgedriver.exe"
DEBUGGER_ADDRESS = "127.0.0.1:9222"
MASTER_XLSX      = r"D:\Kaam\Muneef 0\Test Automation\Weekly Sales summary (week 37).xlsx"
PREV_WEEK_LABEL  = "Sep 01 - Sep 07"  # find this in column A
NEW_WEEK_LABEL   = "Sep 08 - Sep 14"      # write this into the row below

# Map dropdown label → sheet name
dropdown_to_sheet = {
    "Karachi Kabab Wala - Queen Street":    "Kababwala - Queen",
    "Pizza Karachi- Eglinton":       "Pizza K Eglinton",
    "Pizza Karachi -Heartland":      "Pizza K Heartland",
    "Karachi Kabab Wala":            "Kababwala",
    "Karachi Food Court":            "Karachi Food Court",
    "Pizza Karachi Downtown TO":     "Queen St.",
    "Pizza Karachi- Highway Karahi": "Highway",
    "Pizza Karachi - Wonderland":    "Jane",
    "Pizza Karachi - Lebovic":       "Lebovic",
    "Pizza Karachi - Ajax":          "Ajax",
    "Pizza Karachi - Markham Rd":    "Markham",
}

# ───── HELPERS ─────
def parse_currency(txt: str) -> float:
    """Convert strings like '$1,234.56' → 1234.56"""
    return float(txt.replace('$','').replace(',','').strip())

# ───── SETUP SELENIUM ─────
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

# ───── LOAD EXCEL ─────
wb = load_workbook(MASTER_XLSX)

try:
    # 1) Open the restaurant dropdown once
    dd_button = wait.until(EC.element_to_be_clickable((
        By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
    )))
    dd_button.click()
    time.sleep(0.5)

    # 2) Iterate through each restaurant
    for label_text, sheet_name in dropdown_to_sheet.items():
        print(f"▶ Processing {label_text} → sheet {sheet_name}")

        # a) Select restaurant
        xpath_label = (
            f"//div[@data-pw='radio-dropdown-selector']"
            f"//label[.//span[text()='{label_text}']]"
        )
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath_label))).click()

        # b) Wait for the Order-Type table to appear
        wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"] tbody'
        )))
        time.sleep(0.5)

        # c) Grab the summary row
        summary_tr = driver.find_element(
            By.CSS_SELECTOR, 'tr[data-pw="report-table-summary-row"]'
        )
        cells = summary_tr.find_elements(By.TAG_NAME, "td")

        # d) Extract the six numeric values
        values = [parse_currency(cells[i].text) for i in range(1, 7)]
        print("   Summary values:", values)

        # e) Locate the row below the PREV_WEEK_LABEL in column A
        ws = wb[sheet_name]
        target_row = None
        for cell in ws['A']:
            if cell.value == PREV_WEEK_LABEL:
                target_row = cell.row + 1
                break
        if not target_row:
            raise RuntimeError(f"Could not find row for '{PREV_WEEK_LABEL}' in sheet '{sheet_name}'")

        # f) Write new week label in column A and values in columns G–L
        ws.cell(row=target_row, column=1, value=NEW_WEEK_LABEL)
        for idx, val in enumerate(values):
            ws.cell(row=target_row, column=7 + idx, value=val)

        # g) Re-open the dropdown for the next restaurant
        dd_button = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
        )))
        dd_button.click()
        time.sleep(0.5)

    # 3) Save workbook
    wb.save(MASTER_XLSX)
    print(f"\n✅ All restaurants updated below '{PREV_WEEK_LABEL}' for '{NEW_WEEK_LABEL}'")

finally:
    driver.quit()
