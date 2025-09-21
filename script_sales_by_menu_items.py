import os
import time
import glob
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ───── CONFIG ─────
# DRIVER_PATH        = r"C:\Users\PC\Downloads\edgedriver_win64\msedgedriver.exe"
DRIVER_PATH        = "msedgedriver.exe"
DEBUGGER_ADDRESS   = "127.0.0.1:9222"
DOWNLOADS_DIR      = r"C:\Users\PC\Downloads"
# TEMPLATE_CHECKER   = r"D:\Kaam\Muneef 0\Test Automation\Weekly item summary (week 18) .xlsx"
NEW_WEEK           = "Sep 08 - Sep 14"
OUTPUT_XLSX        = fr"D:\Kaam\Muneef 0\Test Automation\{NEW_WEEK}.xlsx"

# Dropdown label → desired sheet name
dropdown_to_sheet = {
    "Karachi Kabab Wala - Queen Street": "Karachi Kabab Wala - Queen",
    "Pizza Karachi- Eglinton":        "Eglinton",
    "Pizza Karachi -Heartland":       "Heartland",
    "Karachi Kabab Wala":             "Karachi Kabab Wala",
    "Karachi Food Court":             "Karachi Food Court",
    "Pizza Karachi Downtown TO":      "Pizza Karachi Downtown TO",
    "Pizza Karachi- Highway Karahi":  "Highway Karahi",
    "Pizza Karachi - Wonderland":     "Wonderland",
    "Pizza Karachi - Lebovic":        "Lebovic",
    "Pizza Karachi - Ajax":           "Ajax",
    "Pizza Karachi - Markham Rd":     "Markham Rd",
}

# ───── HELPERS ─────
def get_latest_file(folder, ext="*.xlsx"):
    files = glob.glob(os.path.join(folder, ext))
    return max(files, key=os.path.getctime)

def download_menu_items_report(driver, wait):
    # Click "Get this report"
    wait.until(EC.element_to_be_clickable((By.XPATH,
        "//button[contains(., 'Get this report')]"
    ))).click()
    time.sleep(0.5)
    # Click "Download"
    wait.until(EC.element_to_be_clickable((By.XPATH,
        "//li[contains(., 'Download')]"
    ))).click()
    time.sleep(0.5)
    # Click "XLSX"
    wait.until(EC.element_to_be_clickable((By.XPATH,
        "//li[contains(., 'XLSX')]"
    ))).click()
    # wait for the file to land
    time.sleep(8)
    return get_latest_file(DOWNLOADS_DIR)

# ───── SETUP SELENIUM ─────
opts = Options()
opts.use_chromium = True
opts.add_experimental_option("debuggerAddress", DEBUGGER_ADDRESS)
driver = webdriver.Edge(service=Service(DRIVER_PATH), options=opts)
wait = WebDriverWait(driver, 20)

# ───── PREPARE NEW WORKBOOK ─────
# Load your template checker sheet
# template_wb      = load_workbook(TEMPLATE_CHECKER, data_only=False)
# template_checker = template_wb["checker"]
# template_checker = template_wb["April 28 - May 04"]

# Create a blank new workbook and drop its default sheet
new_wb = openpyxl.Workbook()
new_wb.remove(new_wb.active)

try:
    # 1) Open the restaurant dropdown once
    dd_btn = wait.until(EC.element_to_be_clickable((
        By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
    )))
    dd_btn.click()
    time.sleep(0.5)

    # 2) For each restaurant: select, download, copy
    for label_text, sheet_name in dropdown_to_sheet.items():
        print(f"▶ Processing {label_text} → sheet {sheet_name}")

        # select its radio button
        xpath = (
            "//div[@data-pw='radio-dropdown-selector']"
            f"//label[.//span[text()='{label_text}']]"
        )
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()

        # wait for the page/table to reload
        wait.until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'table[data-pw="report-table-data"]'
        )))
        time.sleep(0.5)

        # download the Sales-By-Menu-Items
        xlsx_path = download_menu_items_report(driver, wait)
        print("  ↓ Downloaded:", os.path.basename(xlsx_path))

        # load and copy into new workbook
        dl_wb = load_workbook(xlsx_path, data_only=False)
        dl_ws = dl_wb.active

        # create new sheet
        target_ws = new_wb.create_sheet(title=sheet_name)
        # for row in dl_ws.iter_rows():
        #     for cell in row:
        #         tgt = target_ws.cell(row=cell.row, column=cell.column)
        #         tgt.value = cell.value
        #         # optionally copy styles:
        #         # tgt._style = cell._style
        for row in dl_ws.iter_rows():
            for cell in row:
                tgt = target_ws.cell(row=cell.row, column=cell.column)
                val = cell.value

                if (
                    cell.row >= 3 and
                    4 <= cell.column <= 15 and
                    isinstance(val, str)
                ):
                    try:
                        clean_val = val.replace(",", "").strip()
                        if clean_val.endswith("%"):
                            val = float(clean_val.rstrip("%")) / 100
                        else:
                            val = float(clean_val)
                    except ValueError:
                        pass

                tgt.value = val

        os.remove(xlsx_path)

        # re-open dropdown for next loop
        wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR, "button[data-testid='radio-dropdown-selector-input']"
        ))).click()
        time.sleep(0.5)

    # 3) Copy over the template 'checker' sheet
    # checker_ws = new_wb.create_sheet(title="checker")
    # for r in range(1, template_checker.max_row + 1):
    #     for c in range(1, template_checker.max_column + 1):
    #         src = template_checker.cell(row=r, column=c)
    #         dst = checker_ws.cell(row=r, column=c)
    #         dst.value = src.value
            # optionally copy styles:
            # dst._style = src._style

    # 4) Save the completed workbook
    new_wb.save(OUTPUT_XLSX)
    print(f"\n✅ Workbook saved as:\n   {OUTPUT_XLSX}")

finally:
    driver.quit()
